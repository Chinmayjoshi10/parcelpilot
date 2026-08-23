"""Spreadsheet -> typed relational rows.

The v1 design inferred the schema at runtime ("works with any Excel"). This
one declares it, and the difference is the difference between a demo and a
product.

The deterministic policy engine reads `orders.booked_at` and
`orders.cancellation_requested_at` by name and does arithmetic on them. A
schema that can silently change shape cannot support a rule that must give the
same answer every time -- an inferred TEXT column where a timestamptz was
expected does not error, it just makes every cancellation-window calculation
wrong.

Universality is preserved where it actually belongs: `COLUMN_MAPPINGS` is a
declaration, so pointing this at another corpus is a mapping edit, not a code
change. Unmapped columns are preserved verbatim in `raw_json`, so nothing is
lost and every load can be audited against its source.

The other thing this file takes seriously is **timezones**. The sheet says
`2026-08-16 09:00` with no offset, and every cancellation window and SLA
countdown is wall-clock sensitive. Those strings are interpreted in the
tenant's civil timezone, not the server's locale -- otherwise the same data
yields different fee decisions depending on where it is deployed.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl

from agentcore.errors import IngestionError
from agentcore.logging import get_logger

log = get_logger(__name__)


@dataclass(frozen=True)
class ColumnSpec:
    """One canonical column and where it comes from."""

    target: str
    #: Source header names, tried in order. A list rather than a string so a
    #: renamed upstream column is a one-line mapping change.
    sources: tuple[str, ...]
    kind: str  # text | int | decimal | bool | timestamp
    required: bool = False
    #: Value to use when the source column is absent or empty.
    #:
    #: Needed for columns the schema declares NOT NULL: passing an explicit
    #: NULL defeats the database default, so a workbook simply missing an
    #: optional column failed the entire load with a constraint violation --
    #: precisely when onboarding a new customer's spreadsheet. Declaring the
    #: meaning of "absent" here keeps that decision with the mapping rather than
    #: hiding it in the schema.
    default: Any = None
    #: Fixed value written alongside, e.g. currency implied by a *_inr header.
    implies: tuple[str, Any] | None = None


@dataclass(frozen=True)
class TableMapping:
    table: str
    sheet: str
    key: tuple[str, ...]
    columns: tuple[ColumnSpec, ...]


#: The declared contract between the supplied workbook and our schema.
COLUMN_MAPPINGS: tuple[TableMapping, ...] = (
    TableMapping(
        table="accounts",
        sheet="accounts",
        key=("tenant_id", "account_id"),
        columns=(
            ColumnSpec("account_id", ("account_id",), "text", required=True),
            ColumnSpec("account_name", ("account_name",), "text", required=True),
            ColumnSpec("plan", ("plan",), "text"),
            ColumnSpec("status", ("status",), "text"),
            ColumnSpec("csm", ("csm",), "text"),
            # The single most important column in the workbook: it is what makes
            # contract override a foreign-key lookup instead of a retrieval guess.
            ColumnSpec("contract_file", ("contract_file",), "text"),
            # NOT NULL in the schema, so absence means False rather than NULL.
            ColumnSpec("premium_support", ("premium_support",), "bool", default=False),
            ColumnSpec("notes", ("notes",), "text"),
        ),
    ),
    TableMapping(
        table="orders",
        sheet="orders",
        key=("tenant_id", "order_id"),
        columns=(
            ColumnSpec("order_id", ("order_id",), "text", required=True),
            ColumnSpec("account_id", ("account_id",), "text", required=True),
            ColumnSpec("carrier", ("carrier",), "text"),
            ColumnSpec("status", ("status",), "text"),
            ColumnSpec("booked_at", ("booked_at",), "timestamp"),
            ColumnSpec("pickup_window_start", ("pickup_window_start",), "timestamp"),
            ColumnSpec("pickup_window_end", ("pickup_window_end",), "timestamp"),
            ColumnSpec("pickup_actual_at", ("pickup_actual_at",), "timestamp"),
            # The header encodes the currency; we split it into a typed amount
            # plus an explicit currency rather than hiding it in a column name.
            ColumnSpec(
                "shipment_fee",
                ("shipment_fee_inr", "shipment_fee"),
                "decimal",
                implies=("currency", "INR"),
            ),
            # Both NOT NULL: an unrecorded fault is "not at fault", which is
            # also the safe reading -- the SOP forbids promising a credit when
            # fault is unknown, and that is decided by the rule engine.
            ColumnSpec("carrier_fault", ("carrier_fault",), "bool", default=False),
            ColumnSpec("customer_fault", ("customer_fault",), "bool", default=False),
            ColumnSpec(
                "cancellation_requested_at", ("cancellation_requested_at",), "timestamp"
            ),
            ColumnSpec("notes", ("notes",), "text"),
        ),
    ),
    TableMapping(
        table="tickets",
        sheet="tickets",
        key=("tenant_id", "ticket_id"),
        columns=(
            ColumnSpec("ticket_id", ("ticket_id",), "text", required=True),
            ColumnSpec("account_id", ("account_id",), "text", required=True),
            ColumnSpec("created_at", ("created_at",), "timestamp"),
            ColumnSpec("status", ("status",), "text"),
            ColumnSpec("subject", ("subject",), "text"),
            ColumnSpec("description", ("description",), "text"),
            ColumnSpec("channel", ("channel",), "text"),
            ColumnSpec("assigned_to", ("assigned_to",), "text"),
            ColumnSpec(
                "last_customer_message_at", ("last_customer_message_at",), "timestamp"
            ),
            ColumnSpec("historical_resolution", ("historical_resolution",), "text"),
        ),
    ),
)

_TIMESTAMP_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
)


@dataclass
class LoadReport:
    table: str
    rows_read: int = 0
    rows_written: int = 0
    skipped: list[str] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.skipped is None:
            self.skipped = []


def load_workbook_rows(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read every sheet into header-keyed dicts, dropping blank rows.

    The supplied workbook has ~980 trailing empty rows per sheet (an artefact
    of however it was generated). Filtering them here keeps every downstream
    row count honest.
    """
    if not path.exists():
        raise IngestionError(f"structured data file not found: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: dict[str, list[dict[str, Any]]] = {}
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                sheets[sheet.title] = []
                continue

            names = [str(h).strip() if h is not None else "" for h in header]
            records: list[dict[str, Any]] = []
            for values in rows:
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                    continue
                records.append(
                    {name: value for name, value in zip(names, values, strict=False) if name}
                )
            sheets[sheet.title] = records
        return sheets
    finally:
        workbook.close()


def coerce(value: Any, kind: str, *, tz: ZoneInfo, column: str) -> Any:
    """Convert one spreadsheet cell to its typed form.

    Raises rather than substituting a default. A silently-defaulted timestamp
    would make a cancellation-window calculation confidently wrong, which is
    the exact failure this architecture exists to prevent.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    if kind == "text":
        return str(value).strip()

    if kind == "bool":
        if isinstance(value, bool):
            return value
        token = str(value).strip().lower()
        if token in {"1", "true", "yes", "y"}:
            return True
        if token in {"0", "false", "no", "n"}:
            return False
        raise IngestionError(f"{column}: cannot read {value!r} as a boolean")

    if kind == "int":
        try:
            return int(str(value).strip())
        except ValueError as exc:
            raise IngestionError(f"{column}: cannot read {value!r} as an integer") from exc

    if kind == "decimal":
        try:
            return Decimal(str(value).strip())
        except (InvalidOperation, ValueError) as exc:
            raise IngestionError(f"{column}: cannot read {value!r} as a number") from exc

    if kind == "timestamp":
        if isinstance(value, datetime):
            # openpyxl may hand back a real datetime; it is naive, and naive
            # means "tenant civil time" throughout this system.
            return value if value.tzinfo else value.replace(tzinfo=tz)
        token = str(value).strip()
        for fmt in _TIMESTAMP_FORMATS:
            try:
                parsed = datetime.strptime(token, fmt)
            except ValueError:
                continue
            return parsed.replace(tzinfo=tz)
        raise IngestionError(f"{column}: cannot read {value!r} as a timestamp")

    raise IngestionError(f"{column}: unknown column kind {kind}")


def build_rows(
    mapping: TableMapping,
    records: list[dict[str, Any]],
    *,
    tenant_id: str,
    timezone: str,
) -> tuple[list[dict[str, Any]], LoadReport]:
    """Map and coerce a sheet into rows ready for insertion."""
    tz = ZoneInfo(timezone)
    report = LoadReport(table=mapping.table, rows_read=len(records))
    rows: list[dict[str, Any]] = []

    for index, record in enumerate(records, start=2):  # +2: header is row 1
        row: dict[str, Any] = {"tenant_id": tenant_id}
        try:
            for spec in mapping.columns:
                source = next((s for s in spec.sources if s in record), None)
                raw = record.get(source) if source else None

                if raw is None and spec.required:
                    raise IngestionError(
                        f"{mapping.sheet} row {index}: required column "
                        f"{spec.target} is missing or empty"
                    )

                value = coerce(
                    raw, spec.kind, tz=tz, column=f"{mapping.sheet}.{spec.target}"
                )
                row[spec.target] = spec.default if value is None else value
                if spec.implies and row[spec.target] is not None:
                    key, implied = spec.implies
                    row[key] = implied
        except IngestionError as exc:
            # One bad row is skipped and reported; it must not abort the load.
            report.skipped.append(str(exc))
            log.warning("row_skipped", sheet=mapping.sheet, row=index, error=str(exc))
            continue

        # Everything from the source row, including unmapped columns, kept for
        # audit. json default=str so dates and Decimals survive.
        row["raw_json"] = json.dumps(record, default=str, ensure_ascii=False)
        rows.append(row)

    report.rows_written = len(rows)
    return rows, report


def upsert_sql(mapping: TableMapping, columns: list[str]) -> str:
    """Build an idempotent INSERT ... ON CONFLICT for the mapped columns.

    Column names come exclusively from `COLUMN_MAPPINGS` -- module constants,
    never spreadsheet content -- so no untrusted string reaches the statement.
    Values are always bound parameters.
    """
    placeholders = ", ".join(f"%({c})s" for c in columns)
    updatable = [c for c in columns if c not in mapping.key]
    assignments = ", ".join(f"{c} = EXCLUDED.{c}" for c in updatable)
    conflict = ", ".join(mapping.key)
    return (
        f"INSERT INTO {mapping.table} ({', '.join(columns)}) "
        f"VALUES ({placeholders}) "
        f"ON CONFLICT ({conflict}) DO UPDATE SET {assignments}, ingested_at = now()"
    )
