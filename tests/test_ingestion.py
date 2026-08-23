"""Ingestion invariants, asserted against the real ingested corpus.

These run against the live `parcelpilot` tenant produced by
`parcelpilot ingest run`, not a fixture. That is deliberate: the properties
worth protecting here are properties of *this* corpus and *these* six
documents, and a synthetic fixture would let a classification regression pass.
"""

from __future__ import annotations

from datetime import UTC
from zoneinfo import ZoneInfo

import pytest

from agentcore.db import engine
from agentcore.db.engine import fetch_all, fetch_one
from agentcore.ingestion.documents import MAX_CHUNK_CHARS, normalise
from agentcore.settings import load_config
from agentcore.types import Principal, Role

TENANT = load_config().tenant.id


@pytest.fixture(scope="module")
def corpus() -> list[dict]:
    with engine.admin() as conn:
        rows = fetch_all(
            conn,
            """
            SELECT d.*, iv.status AS version_status
            FROM documents d
            JOIN index_versions iv USING (index_version_id)
            WHERE d.tenant_id = %s AND iv.status = 'active'
            """,
            (TENANT,),
        )
    if not rows:
        pytest.skip("run `parcelpilot ingest run` before this suite")
    return rows


def _by_name(corpus: list[dict], fragment: str) -> dict:
    return next(d for d in corpus if fragment in d["filename"])


class TestTrustModel:
    def test_deprecated_policy_cannot_ground_an_answer(self, corpus):
        """The central trust property.

        v2 is not merely ranked lower -- its eligibility forbids grounding. A
        strong lexical match against it can never produce a citation, which is
        exactly what a relevance-times-authority score would have allowed.
        """
        v2 = _by_name(corpus, "02_Support_Policy_v2_DEPRECATED")
        assert v2["source_class"] == "policy_deprecated"
        assert v2["eligibility"] == "conflict_only"
        assert v2["freshness"] == "deprecated"
        assert v2["authority"] == 0

    def test_current_policy_outranks_nothing_by_accident(self, corpus):
        v3 = _by_name(corpus, "01_Support_Policy_v3_CURRENT")
        assert v3["eligibility"] == "groundable"
        assert v3["freshness"] == "current"

    def test_versions_of_one_policy_share_a_family(self, corpus):
        """Makes v2-vs-v3 conflict detection a join rather than filename parsing."""
        family = [d for d in corpus if d["policy_family"] == "support_policy"]
        assert {d["version_label"] for d in family} == {"v2", "v3"}
        current = [d for d in family if d["freshness"] == "current"]
        assert len(current) == 1, "exactly one version of a policy may be in force"

    def test_contracts_outrank_general_policy(self, corpus):
        contracts = [d for d in corpus if d["source_class"] == "customer_agreement"]
        policies = [d for d in corpus if d["source_class"] == "policy_current"]
        assert contracts and policies
        assert min(c["authority"] for c in contracts) > max(p["authority"] for p in policies)

    def test_every_document_resolved_to_a_known_class(self, corpus):
        """Nothing fell through to the context_only fail-safe.

        If a real policy document ever lands as ticket_resolution, it silently
        stops being able to answer anything -- so assert the fallback is unused
        on a corpus we understand.
        """
        assert [d["filename"] for d in corpus if d["source_class"] == "ticket_resolution"] == []


class TestContractOwnership:
    def test_contracts_are_bound_to_their_account(self, corpus):
        assert _by_name(corpus, "05_Northstar")["owner_account_id"] == "ACCT-001"
        assert _by_name(corpus, "06_LumenWorks")["owner_account_id"] == "ACCT-002"

    def test_general_sources_are_not_account_scoped(self, corpus):
        for fragment in ("01_Support_Policy_v3", "03_Cancellation", "04_Product"):
            assert _by_name(corpus, fragment)["owner_account_id"] is None

    def test_contract_file_foreign_key_matches_ingested_owner(self):
        """The lookup that makes contract override deterministic.

        If these two ever disagree the override is applied to the wrong
        account, so ingestion refuses -- this asserts they agree.
        """
        with engine.admin() as conn:
            rows = fetch_all(
                conn,
                """
                SELECT a.account_id, a.contract_file, d.owner_account_id
                FROM accounts a
                JOIN documents d ON d.filename = a.contract_file
                JOIN index_versions iv USING (index_version_id)
                WHERE a.tenant_id = %s
                  AND a.contract_file IS NOT NULL
                  AND iv.status = 'active'
                """,
                (TENANT,),
            )
        # Scoped to the active version on purpose: superseded versions retain
        # their own document rows (that is what makes rollback possible), so an
        # unscoped join multiplies by the number of ingests ever run.
        assert len(rows) == 2
        for row in rows:
            assert row["account_id"] == row["owner_account_id"]


class TestContractIsolation:
    """A customer must not be able to read another customer's contract.

    This is the highest-consequence leak in the system: contracts contain
    negotiated commercial terms. Enforced by RLS on `documents` and `chunks`
    via the denormalised `owner_account_id`.
    """

    def test_customer_sees_own_contract_and_general_sources(self):
        principal = Principal(
            tenant_id=TENANT, user_id="u", role=Role.CUSTOMER, account_id="ACCT-001"
        )
        with engine.scoped(principal, read_only=True) as conn:
            visible = {r["filename"] for r in fetch_all(conn, "SELECT filename FROM documents")}

        assert "05_Northstar_Logistics_Enterprise_Agreement.pdf" in visible
        assert "01_Support_Policy_v3_CURRENT.pdf" in visible

    def test_customer_cannot_see_another_customers_contract(self):
        principal = Principal(
            tenant_id=TENANT, user_id="u", role=Role.CUSTOMER, account_id="ACCT-001"
        )
        with engine.scoped(principal, read_only=True) as conn:
            visible = {r["filename"] for r in fetch_all(conn, "SELECT filename FROM documents")}
            # Unfiltered chunk read: if RLS were misconfigured this would leak
            # the text of the other agreement, not just its existence.
            chunk_owners = {
                r["owner_account_id"]
                for r in fetch_all(conn, "SELECT owner_account_id FROM chunks")
            }

        assert "06_LumenWorks_Service_Agreement.pdf" not in visible
        assert chunk_owners <= {None, "ACCT-001"}

    def test_account_without_a_contract_sees_only_general_sources(self):
        """ACCT-003 has no agreement, so it must see no contract at all."""
        principal = Principal(
            tenant_id=TENANT, user_id="u", role=Role.CUSTOMER, account_id="ACCT-003"
        )
        with engine.scoped(principal, read_only=True) as conn:
            classes = {
                r["source_class"] for r in fetch_all(conn, "SELECT source_class FROM documents")
            }
        assert "customer_agreement" not in classes

    def test_staff_see_every_contract_in_the_tenant(self):
        principal = Principal(tenant_id=TENANT, user_id="ops", role=Role.OPERATIONS_ADMIN)
        with engine.scoped(principal, read_only=True) as conn:
            owners = {
                r["owner_account_id"]
                for r in fetch_all(
                    conn,
                    "SELECT owner_account_id FROM documents WHERE owner_account_id IS NOT NULL",
                )
            }
        assert owners == {"ACCT-001", "ACCT-002"}


class TestStructuredData:
    def test_expected_row_counts(self):
        """The workbook carries ~980 blank trailing rows per sheet.

        Loading them would inflate every count and every dashboard aggregate.
        """
        with engine.admin() as conn:
            counts = {
                table: fetch_one(
                    conn, f"SELECT count(*) AS n FROM {table} WHERE tenant_id = %s", (TENANT,)
                )["n"]
                for table in ("accounts", "orders", "tickets")
            }
        assert counts == {"accounts": 4, "orders": 6, "tickets": 7}

    def test_timestamps_are_in_tenant_civil_time(self):
        """`2026-08-16 09:00` in the sheet means 09:00 in Asia/Kolkata.

        Interpreted as UTC instead, every cancellation-window and SLA
        calculation shifts by 5h30m -- which silently changes fee decisions.
        """
        with engine.admin() as conn:
            row = fetch_one(
                conn,
                "SELECT booked_at FROM orders WHERE tenant_id = %s AND order_id = 'ORD-1001'",
                (TENANT,),
            )
        booked = row["booked_at"].astimezone(ZoneInfo("Asia/Kolkata"))
        assert (booked.hour, booked.minute) == (9, 0)
        # 09:00 IST is 03:30 UTC. Asserted explicitly so a regression to naive
        # UTC parsing fails here rather than in a fee calculation.
        assert booked.astimezone(UTC).hour == 3

    def test_numeric_and_boolean_columns_are_typed(self):
        """The policy engine does arithmetic and boolean logic on these.

        A TEXT column holding '4200.0' and '1' would not error -- it would just
        make every comparison wrong.
        """
        with engine.admin() as conn:
            row = fetch_one(
                conn,
                """
                SELECT shipment_fee, currency, carrier_fault, customer_fault, status
                FROM orders WHERE tenant_id = %s AND order_id = 'ORD-2002'
                """,
                (TENANT,),
            )
        assert float(row["shipment_fee"]) == 2400.0
        assert row["currency"] == "INR"
        assert row["carrier_fault"] is True
        assert row["customer_fault"] is False

    def test_unmapped_source_columns_are_preserved(self):
        with engine.admin() as conn:
            row = fetch_one(
                conn,
                "SELECT raw_json FROM orders WHERE tenant_id = %s AND order_id = 'ORD-1001'",
                (TENANT,),
            )
        # The source header is shipment_fee_inr; we split it into a typed
        # amount plus a currency, and the original is still auditable.
        assert "shipment_fee_inr" in row["raw_json"]

    def test_historical_resolutions_are_loaded_but_never_authoritative(self):
        """TKT-450 records a wrong answer a human gave.

        It must be present -- the system has to be able to notice and flag the
        discrepancy -- while never being able to ground a policy claim. That is
        enforced by ticket_resolution being context_only in config.yaml.
        """
        with engine.admin() as conn:
            row = fetch_one(
                conn,
                """
                SELECT historical_resolution FROM tickets
                WHERE tenant_id = %s AND ticket_id = 'TKT-450'
                """,
                (TENANT,),
            )
        assert row["historical_resolution"] is not None
        assert "250" in row["historical_resolution"]

        sources = load_config().sources
        from agentcore.types import SourceClass

        assert SourceClass.TICKET_RESOLUTION not in sources.groundable


class TestChunking:
    def test_chunks_align_to_numbered_sections(self):
        """Section-aligned chunks are what make a citation land on a clause.

        The SOP's cancellation rules and its credit rules must not share a
        chunk, or a citation cannot distinguish which rule it relied on.
        """
        with engine.admin() as conn:
            rows = fetch_all(
                conn,
                """
                SELECT c.section_path, c.text FROM chunks c
                JOIN documents d USING (document_id)
                WHERE d.filename = '03_Cancellation_and_Service_Credit_SOP_v4.pdf'
                ORDER BY c.ordinal
                """,
            )
        headings = [r["section_path"] for r in rows if r["section_path"]]
        assert any("cancellation" in h.lower() for h in headings)
        assert any("credit" in h.lower() for h in headings)

        cancellation = next(r for r in rows if r["section_path"] and "1." in r["section_path"])
        assert "INR 250" in cancellation["text"]

    def test_no_chunk_exceeds_the_size_ceiling(self):
        with engine.admin() as conn:
            row = fetch_one(conn, "SELECT max(length(text)) AS longest FROM chunks")
        assert row["longest"] <= MAX_CHUNK_CHARS

    def test_full_text_search_vector_is_populated(self):
        """Lexical retrieval is the half that works without an embedder."""
        with engine.admin() as conn:
            row = fetch_one(
                conn, "SELECT count(*) AS n FROM chunks WHERE tsv IS NULL OR tsv = ''::tsvector"
            )
        assert row["n"] == 0

    def test_invisible_characters_are_stripped(self):
        """These PDFs bullet with U+25CF followed by U+200B.

        Left in place, a quote copied from a bullet would never match the
        stored text, so the citation validator would reject valid citations.
        The bullet marker is also normalised to exactly one trailing space, so
        a bullet written without one does not produce "-DRAFT" beside
        "- BOOKED" in the same list.
        """
        assert normalise("●​DRAFT: May be cancelled") == "- DRAFT: May be cancelled"
        # Idempotent: normalising twice must not drift, or re-ingestion would
        # invalidate every citation issued against the previous index version.
        once = normalise("●​ DRAFT: no fee.")
        assert normalise(once) == once

        # Escaped and bound rather than embedded literally: a raw U+200B inside
        # a source file is invisible to review and encoding-fragile on Windows.
        with engine.admin() as conn:
            row = fetch_one(
                conn,
                "SELECT count(*) AS n FROM chunks WHERE text LIKE %s",
                (f"%{chr(0x200B)}%",),
            )
        assert row["n"] == 0


class TestIndexVersioning:
    def test_exactly_one_active_version(self):
        """Enforced by a unique partial index, so a bad flip is rejected."""
        with engine.admin() as conn:
            row = fetch_one(
                conn,
                """
                SELECT count(*) AS n FROM index_versions
                WHERE tenant_id = %s AND status = 'active'
                """,
                (TENANT,),
            )
        assert row["n"] == 1

    def test_active_version_counts_match_reality(self):
        """Guards against a version activated while half-built."""
        with engine.admin() as conn:
            version = fetch_one(
                conn,
                """
                SELECT index_version_id, document_count, chunk_count
                FROM index_versions WHERE tenant_id = %s AND status = 'active'
                """,
                (TENANT,),
            )
            actual = fetch_one(
                conn,
                """
                SELECT (SELECT count(*) FROM documents WHERE index_version_id = %(v)s) AS docs,
                       (SELECT count(*) FROM chunks WHERE index_version_id = %(v)s) AS chunks
                """,
                {"v": version["index_version_id"]},
            )
        assert version["document_count"] == actual["docs"] > 0
        assert version["chunk_count"] == actual["chunks"] > 0


class TestUnmappedSheets:
    """A new company's workbook must not fail silently.

    The loader previously `continue`d past any sheet no mapping claimed, which
    on a new corpus meant zero rows, no error, and an ingest that looked like it
    worked. The mapping stays declared rather than inferred -- the policy engine
    reads column names -- but the boundary is now visible.
    """

    def test_unmapped_sheets_are_reported(self, tmp_path):

        import openpyxl

        from agentcore.ingestion.pipeline import IngestReport, _load_structured

        workbook = openpyxl.Workbook()
        # A sheet no mapping claims, carrying real rows.
        shipments = workbook.active
        shipments.title = "shipments"
        shipments.append(["shipment_ref", "weight_kg", "destination"])
        shipments.append(["SHP-1", 12.5, "Berlin"])
        shipments.append(["SHP-2", 3.0, "Paris"])
        # Documentation, which is expected to be skipped without complaint.
        readme = workbook.create_sheet("README")
        readme.append(["This workbook is documentation."])
        workbook.save(tmp_path / "other_company.xlsx")

        config = load_config().model_copy(
            update={
                "data": load_config().data.model_copy(
                    update={"structured_dir": tmp_path}
                )
            }
        )
        report = IngestReport()

        with engine.admin() as conn:
            # No mapped sheet is present, so nothing is written; we are
            # asserting on what gets REPORTED.
            _load_structured(conn, config, report)
            conn.rollback()

        sheets = {entry["sheet"] for entry in report.unmapped_sheets}
        assert "shipments" in sheets
        # README is documentation, not data: skipped quietly on purpose.
        assert "README" not in sheets

        entry = next(e for e in report.unmapped_sheets if e["sheet"] == "shipments")
        assert entry["rows"] == 2
        assert "shipment_ref" in entry["columns"]

        # And it reaches the warnings the CLI prints, not just a log line.
        assert any("shipments" in w for w in report.warnings)
        assert any("no column mapping claims them" in w for w in report.warnings)

    def test_absent_mapped_sheets_are_reported(self, tmp_path):
        """The other direction: a workbook missing a sheet we expect."""
        import openpyxl

        from agentcore.ingestion.pipeline import IngestReport, _load_structured

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "accounts"
        sheet.append(["account_id", "account_name"])
        sheet.append(["ACCT-999", "Only Accounts Ltd"])
        workbook.save(tmp_path / "partial.xlsx")

        base = load_config()
        config = base.model_copy(
            update={"data": base.data.model_copy(update={"structured_dir": tmp_path})}
        )
        report = IngestReport()

        with engine.admin() as conn:
            _load_structured(conn, config, report)
            conn.rollback()

        assert any("orders" in w and "absent" in w for w in report.warnings)
